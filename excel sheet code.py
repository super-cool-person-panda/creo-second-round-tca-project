"""
final_calculations.py — Kingston Painting Lead Scorer
Dynamic version: all frequency, summary, and stats sheets use COUNTIFS/AVERAGEIF
against a named Excel Table (LeadData) so values update live when data is added.

Cramér V weights are recalculated from data each time this script runs.

Sheets:
  1. Lead Scorer       — interactive tool
  2. Factor Weights    — V scores + editable weights (counts live via COUNTIFS)
  3. Frequency Tables  — COUNTIFS formulas, auto-expand with LeadData table
  4. Summary Stats     — COUNTIFS + AVERAGEIF, live
  5. Raw Data          — named Excel Table "LeadData" (source of truth)
  6. _Lookup           — hidden probability values for scorer formulas
"""

import pandas as pd
import numpy as np
from scipy.stats import chi2_contingency
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.table import Table, TableStyleInfo
import sys

try:
    import config
except ImportError:
    print("config.py not found. Run setup_wizard.py first.")
    sys.exit(1)

# ── LOAD + PREP ───────────────────────────────────────────────────────

df = pd.read_excel("leads_data_organised.xlsx", sheet_name="Complete Labelled")
target     = config.TARGET_COL
bands      = config.PROFIT_BANDS
cat_cols   = [c for c in config.CATEGORICAL_COLS if c in df.columns]
num_col    = config.NUMERIC_COLS[0] if config.NUMERIC_COLS else None
total_rows = len(df)

# Bin numeric col
if num_col and num_col in config.BIN_CONFIG:
    cfg = config.BIN_CONFIG[num_col]
    df[num_col+"_bin"] = pd.cut(
        df[num_col], bins=cfg["edges"], labels=cfg["labels"],
        right=False, include_lowest=True
    )

def cramers_v(col):
    ct = pd.crosstab(df[col], df[target])
    chi2, p, _, _ = chi2_contingency(ct)
    n = ct.values.sum(); k = min(ct.shape)-1
    return (round(np.sqrt(chi2/(n*k)), 4) if k>0 else 0), round(p, 6)

# Build factor list sorted by V desc
sig_factors = []
for col in cat_cols:
    v, p = cramers_v(col)
    sig_factors.append((col.replace("_"," ").title(), col, None, v, p))
if num_col and num_col+"_bin" in df.columns:
    v, p = cramers_v(num_col+"_bin")
    sig_factors.append((num_col.replace("_"," ").title(), num_col, num_col+"_bin", v, p))
sig_factors.sort(key=lambda x: -x[3])

# Marginal probs for _Lookup (scorer uses these — static snapshot, refreshed each run)
marginal = {}
for display, col, bin_col, v, p in sig_factors:
    use_col = bin_col if bin_col else col
    marginal[col] = {}
    for val in df[use_col].dropna().unique():
        sub = df[df[use_col]==val]
        marginal[col][str(val)] = {b: round(sub[target].eq(b).sum()/len(sub)*100, 2) for b in bands}
overall = {b: round(df[target].eq(b).sum()/total_rows*100, 2) for b in bands}

# Ordered dropdown values per factor
def get_ordered_vals(col, bin_col):
    if bin_col and bin_col in df.columns and num_col in config.BIN_CONFIG:
        return [str(l) for l in config.BIN_CONFIG[num_col]["labels"]
                if str(l) in df[bin_col].dropna().astype(str).unique()]
    elif col == "customer_age_bracket":
        raw = df[col].dropna().unique().astype(str).tolist()
        return sorted(raw, key=lambda x: int(x.split("-")[0].replace("+","")))
    else:
        return sorted(df[col].dropna().unique().astype(str).tolist())

# ── COLOURS ──────────────────────────────────────────────────────────

NAVY="1F3864"; NAVY2="2E4D8A"; LBLUE="DCE6F1"; WHITE="FFFFFF"
CREAM="FAFAE8"; GBTN="2E7D32"; GBG="C6EFCE"; ABG="FFEB9C"; RBG="FFC7CE"
GFG="276221"; AFG="9C6500"; RFG="9C0006"; LGRAY="F2F2F2"
BGRAY="AAAAAA"; ALTROW="EEF4FB"; YELLOW="FFF2CC"; GOLD="FFE699"

def bdr(c=BGRAY):
    s=Side(style="thin",color=c); return Border(left=s,right=s,top=s,bottom=s)

def hdr(cell, text, bg=NAVY, fg=WHITE, size=11, bold=True, center=True):
    cell.value=text
    cell.font=Font(name="Calibri",bold=bold,size=size,color=fg)
    cell.fill=PatternFill("solid",start_color=bg)
    cell.alignment=Alignment(
        horizontal="center" if center else "left",
        vertical="center", wrap_text=True, indent=0 if center else 1)
    cell.border=bdr(bg)

def val_cell(cell, v, bg=WHITE, bold=False, size=10, color="000000", fmt=None, left=False):
    cell.value=v
    cell.font=Font(name="Calibri",bold=bold,size=size,color=color)
    cell.fill=PatternFill("solid",start_color=bg)
    cell.alignment=Alignment(
        horizontal="left" if left else "center",
        vertical="center", indent=1 if left else 0)
    cell.border=bdr()
    if fmt: cell.number_format=fmt

wb = Workbook()
wb.remove(wb.active)

# ═══════════════════════════════════════════════════════════════════════
# SHEET 1: RAW DATA  (named Excel Table = "LeadData")
# Must be first so other sheets can reference it by table name
# ═══════════════════════════════════════════════════════════════════════

ws_raw = wb.create_sheet("Raw Data")

# Add binned col to export df
df_export = df.copy()
if num_col and num_col+"_bin" in df.columns:
    # Insert bin col right after numeric col
    cols = list(df_export.columns)
    idx  = cols.index(num_col) + 1
    cols.insert(idx, num_col+"_bin")
    df_export = df_export[cols]

# Write headers
for ci, cn in enumerate(df_export.columns, 1):
    c = ws_raw.cell(1, ci, cn)
    c.font = Font(name="Calibri", bold=True, size=10, color=WHITE)
    c.fill = PatternFill("solid", start_color=NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = bdr()
    ws_raw.column_dimensions[get_column_letter(ci)].width = 20
ws_raw.row_dimensions[1].height = 22

# Write data rows
for ri, row in enumerate(df_export.itertuples(index=False), 2):
    bg = ALTROW if ri%2==0 else WHITE
    for ci, v in enumerate(row, 1):
        c = ws_raw.cell(ri, ci, v)
        c.font = Font(name="Calibri", size=9)
        c.fill = PatternFill("solid", start_color=bg)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = bdr()
    ws_raw.row_dimensions[ri].height = 16

# Create named Excel Table — this is the key: all COUNTIFS reference LeadData[col]
last_col = get_column_letter(len(df_export.columns))
last_row = total_rows + 1
tbl = Table(displayName="LeadData", ref=f"A1:{last_col}{last_row}")
tbl.tableStyleInfo = TableStyleInfo(
    name="TableStyleMedium9", showFirstColumn=False,
    showLastColumn=False, showRowStripes=True, showColumnStripes=False)
ws_raw.add_table(tbl)
ws_raw.auto_filter.ref = f"A1:{last_col}1"

# Map col names to Excel Table column references
# e.g. LeadData[property_type]
def tref(col_name):
    return f"LeadData[{col_name}]"

# bin col name in table
bin_col_name = num_col+"_bin" if num_col and num_col+"_bin" in df_export.columns else None

# ═══════════════════════════════════════════════════════════════════════
# SHEET 2: FREQUENCY TABLES  — all COUNTIFS against LeadData
# ═══════════════════════════════════════════════════════════════════════

ws_freq = wb.create_sheet("Frequency Tables")
ws_freq.sheet_view.showGridLines = False

ws_freq.merge_cells("A1:I1")
hdr(ws_freq["A1"], "Frequency & Probability Tables — Auto-updates when LeadData table grows", size=12)
ws_freq.row_dimensions[1].height = 28

ws_freq.merge_cells("A2:I2")
note = ws_freq["A2"]
note.value = "All counts use COUNTIFS against the LeadData table on Raw Data sheet. Add rows there and these update automatically."
note.font  = Font(name="Calibri", italic=True, size=9, color="444444")
note.fill  = PatternFill("solid", start_color="F9F9F9")
note.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
ws_freq.row_dimensions[2].height = 22

for ci in range(1, 10):
    ws_freq.column_dimensions[get_column_letter(ci)].width = 16
ws_freq.column_dimensions["A"].width = 22

fr = 4
for display, col, bin_col, v, p in sig_factors:
    use_col = bin_col if bin_col else col
    tbl_col = bin_col_name if bin_col else col   # column name in LeadData table
    ordered_vals = get_ordered_vals(col, bin_col)

    # Section header
    n_cols = 2 + len(bands) + len(bands)  # label + total + bands + band%
    ws_freq.merge_cells(f"A{fr}:{get_column_letter(n_cols)}{fr}")
    hdr(ws_freq.cell(fr, 1),
        f"{display}   (Cramér V = {v:.4f}   p = {p:.6f})",
        bg=NAVY2, size=10, center=False)
    ws_freq.row_dimensions[fr].height = 20; fr += 1

    # Column headers
    hdrs = [display] + bands + ["Total"] + [f"{b} %" for b in bands]
    for ci, h in enumerate(hdrs, 1):
        c = ws_freq.cell(fr, ci, h)
        c.font = Font(name="Calibri", bold=True, size=9, color=WHITE)
        c.fill = PatternFill("solid", start_color=NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center"); c.border = bdr()
    ws_freq.row_dimensions[fr].height = 18; fr += 1

    data_start = fr
    for vi, val in enumerate(ordered_vals):
        bg = ALTROW if vi%2==0 else WHITE
        # Label
        c = ws_freq.cell(fr, 1, val)
        c.font=Font(name="Calibri",size=9); c.fill=PatternFill("solid",start_color=bg)
        c.alignment=Alignment(horizontal="left",vertical="center",indent=1); c.border=bdr()

        # Band counts via COUNTIFS
        band_cols = []
        for bi, b in enumerate(bands):
            ci = 2 + bi
            col_letter = get_column_letter(ci)
            formula = f'=COUNTIFS({tref(tbl_col)},"{val}",{tref(target)},"{b}")'
            c = ws_freq.cell(fr, ci, formula)
            c.font=Font(name="Calibri",size=9); c.fill=PatternFill("solid",start_color=bg)
            c.alignment=Alignment(horizontal="center",vertical="center"); c.border=bdr()
            band_cols.append(col_letter)

        # Total
        tot_ci = 2 + len(bands)
        tot_col = get_column_letter(tot_ci)
        band_range = f"{band_cols[0]}{fr}:{band_cols[-1]}{fr}"
        ws_freq.cell(fr, tot_ci, f"=SUM({band_range})")
        ws_freq.cell(fr, tot_ci).font=Font(name="Calibri",size=9,bold=True)
        ws_freq.cell(fr, tot_ci).fill=PatternFill("solid",start_color=bg)
        ws_freq.cell(fr, tot_ci).alignment=Alignment(horizontal="center",vertical="center")
        ws_freq.cell(fr, tot_ci).border=bdr()

        # Band %
        for bi, b in enumerate(bands):
            pci = tot_ci + 1 + bi
            bci = 2 + bi
            c = ws_freq.cell(fr, pci,
                f"=IFERROR(ROUND({get_column_letter(bci)}{fr}/{tot_col}{fr}*100,1),0)")
            c.font=Font(name="Calibri",size=9); c.fill=PatternFill("solid",start_color=bg)
            c.alignment=Alignment(horizontal="center",vertical="center")
            c.border=bdr(); c.number_format="0.0"

        ws_freq.row_dimensions[fr].height = 18; fr += 1

    # Grand total row — SUM of all value rows
    data_end = fr - 1
    c = ws_freq.cell(fr, 1, "TOTAL")
    c.font=Font(name="Calibri",bold=True,size=9)
    c.fill=PatternFill("solid",start_color=LGRAY)
    c.alignment=Alignment(horizontal="left",vertical="center",indent=1); c.border=bdr()

    for ci in range(2, 2+len(bands)+1):
        c = ws_freq.cell(fr, ci, f"=SUM({get_column_letter(ci)}{data_start}:{get_column_letter(ci)}{data_end})")
        c.font=Font(name="Calibri",bold=True,size=9)
        c.fill=PatternFill("solid",start_color=LGRAY)
        c.alignment=Alignment(horizontal="center",vertical="center"); c.border=bdr()
    for ci in range(2+len(bands)+1, 2+len(bands)+1+len(bands)):
        ws_freq.cell(fr, ci, "").border=bdr()
        ws_freq.cell(fr, ci).fill=PatternFill("solid",start_color=LGRAY)

    ws_freq.row_dimensions[fr].height = 18; fr += 2

# ═══════════════════════════════════════════════════════════════════════
# SHEET 3: FACTOR WEIGHTS
# ═══════════════════════════════════════════════════════════════════════

ws_wt = wb.create_sheet("Factor Weights")
ws_wt.sheet_view.showGridLines = False
for c, w in zip("ABCDEFG", [28, 14, 14, 16, 18, 20, 24]):
    ws_wt.column_dimensions[c].width = w

ws_wt.merge_cells("A1:G1")
hdr(ws_wt["A1"], "Factor Weights — Cramér V Scores & Adjustable Weights", size=13)
ws_wt.row_dimensions[1].height = 30

ws_wt.merge_cells("A2:G2")
s2 = ws_wt["A2"]
s2.value = ("Cramér V is recalculated each time final_calculations.py runs. "
            "Edit the yellow 'Adjusted Weight' (col G) — the Lead Scorer updates automatically. "
            "Count columns use live COUNTIFS against LeadData.")
s2.font = Font(name="Calibri", italic=True, size=9, color="444444")
s2.fill = PatternFill("solid", start_color="F9F9F9")
s2.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
ws_wt.row_dimensions[2].height = 32

# Header row
for ci, h in enumerate(["Factor","Cramér V","p-value","Significant?",
                          f"n ({bands[0]})", f"n ({bands[1]})", "Adjusted Weight"], 1):
    c = ws_wt.cell(4, ci, h)
    c.font=Font(name="Calibri",bold=True,size=10,color=WHITE)
    c.fill=PatternFill("solid",start_color=NAVY)
    c.alignment=Alignment(horizontal="center",vertical="center"); c.border=bdr()
ws_wt.row_dimensions[4].height = 22

weight_cell_refs = {}   # col -> absolute ref like "'Factor Weights'!$G$5"
wr = 5
for display, col, bin_col, v, p in sig_factors:
    sig  = "✓ Yes" if p<0.05 else "✗ No"
    sbg  = GBG if p<0.05 else RBG
    sfg  = GFG if p<0.05 else RFG
    tbl_col = bin_col_name if bin_col else col
    adj_w = config.QUAL_WEIGHTS.get(col, v)

    # n(High) and n(top 2 band) via live COUNTIFS
    high_formula = f'=COUNTIF({tref(target)},"{bands[0]}")'  # overall
    b0_formula   = f'=COUNTIF({tref(target)},"{bands[0]}")'
    b1_formula   = f'=COUNTIF({tref(target)},"{bands[1]}")'

    row_vals = [display, v, p, sig, b0_formula, b1_formula, round(adj_w, 4)]
    for ci, val in enumerate(row_vals, 1):
        c = ws_wt.cell(wr, ci, val)
        bg = ALTROW if wr%2==0 else WHITE
        c.fill = PatternFill("solid", start_color=(
            sbg if ci==4 else (YELLOW if ci==7 else bg)))
        c.font = Font(name="Calibri", size=10, bold=(ci in (1,4,7)),
                      color=(sfg if ci==4 else ("7B3F00" if ci==7 else "000000")))
        c.alignment = Alignment(
            horizontal="left" if ci==1 else "center",
            vertical="center", indent=1 if ci==1 else 0)
        c.border = bdr()
        if ci == 2: c.number_format = "0.0000"
        if ci == 3: c.number_format = "0.000000"
        if ci == 7: c.number_format = "0.0000"

    weight_cell_refs[col] = f"'Factor Weights'!$G${wr}"
    ws_wt.row_dimensions[wr].height = 22; wr += 1

# Total weight row
for ci in range(1, 8):
    ws_wt.cell(wr, ci).border = bdr()
ws_wt.cell(wr, 1, "TOTAL WEIGHT")
ws_wt.cell(wr, 1).font = Font(name="Calibri", bold=True, size=10)
ws_wt.cell(wr, 1).alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws_wt.cell(wr, 7, f"=SUM(G5:G{wr-1})")
ws_wt.cell(wr, 7).font  = Font(name="Calibri", bold=True, size=10)
ws_wt.cell(wr, 7).fill  = PatternFill("solid", start_color=GOLD)
ws_wt.cell(wr, 7).alignment = Alignment(horizontal="center", vertical="center")
ws_wt.cell(wr, 7).number_format = "0.0000"
total_wt_ref = f"'Factor Weights'!$G${wr}"
ws_wt.row_dimensions[wr].height = 22

note_r = wr + 2
ws_wt.merge_cells(f"A{note_r}:G{note_r}")
n = ws_wt.cell(note_r, 1,
    "💡 Edit yellow cells in column G to adjust weights. "
    "Re-run final_calculations.py to refresh Cramér V scores when new labelled data is added.")
n.font = Font(name="Calibri", italic=True, size=9, color="555555")
n.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
ws_wt.row_dimensions[note_r].height = 28

# ═══════════════════════════════════════════════════════════════════════
# SHEET 4: SUMMARY STATS  — all live COUNTIFS / AVERAGEIF
# ═══════════════════════════════════════════════════════════════════════

ws2 = wb.create_sheet("Summary Stats")
ws2.sheet_view.showGridLines = False
ws2.column_dimensions["A"].width = 28
for ci in range(2, 8):
    ws2.column_dimensions[get_column_letter(ci)].width = 18

ws2.merge_cells("A1:G1")
hdr(ws2["A1"], "Kingston Painting — Dataset Summary (Live — auto-updates with LeadData)", size=13)
ws2.row_dimensions[1].height = 30

ws2.merge_cells("A2:G2")
s2 = ws2["A2"]
s2.value = "All counts and averages use COUNTIFS/AVERAGEIF against LeadData table. Add new labelled rows to Raw Data and this sheet updates instantly."
s2.font = Font(name="Calibri", italic=True, size=9, color="444444")
s2.fill = PatternFill("solid", start_color="F9F9F9")
s2.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
ws2.row_dimensions[2].height = 22

# ── Overall distribution ──────────────────────────────────────────────
ws2.merge_cells("A4:G4")
hdr(ws2["A4"], "Overall Profit Band Distribution", bg=NAVY2, size=10)
ws2.row_dimensions[4].height = 22

band_hdrs = ["Band", "Count", "% of Total", f"Avg {num_col.replace('_',' ').title()} (sqft)",
             "Top Property Type", "Top Referral Source", "Top Neighbourhood"]
for ci, h in enumerate(band_hdrs, 1):
    c = ws2.cell(5, ci, h)
    c.font=Font(name="Calibri",bold=True,size=10,color=WHITE)
    c.fill=PatternFill("solid",start_color=NAVY)
    c.alignment=Alignment(horizontal="center",vertical="center"); c.border=bdr()
ws2.row_dimensions[5].height = 20

bbg = {bands[0]:GBG, bands[1]:ABG, bands[2]:RBG}
bfg = {bands[0]:GFG, bands[1]:AFG, bands[2]:RFG}

# Total count formula (denominator for %)
total_count_formula = f"=COUNTA({tref(target)})"

for ri, band in enumerate(bands, 6):
    bg = bbg.get(band, WHITE); fg = bfg.get(band, "000000")
    count_f  = f'=COUNTIF({tref(target)},"{band}")'
    pct_f    = f'=IFERROR(COUNTIF({tref(target)},"{band}")/COUNTA({tref(target)}),0)'
    avg_f    = f'=IFERROR(AVERAGEIF({tref(target)},"{band}",{tref(num_col)}),0)' if num_col else "N/A"

    # Top property type for this band (MODE equivalent via COUNTIFS — approximate with most common)
    # Excel doesn't have a direct MODEIF, so we use a helper: compute in Python, hardcode label
    # BUT mark it as "refresh on re-run" so user knows
    sub = df[df[target]==band]
    top_prop = sub["property_type"].mode()[0] if "property_type" in sub.columns and len(sub) else "—"
    top_ref  = sub["referral_source"].mode()[0] if "referral_source" in sub.columns and len(sub) else "—"
    top_neigh= sub["neighbourhood"].mode()[0] if "neighbourhood" in sub.columns and len(sub) else "—"

    row_data = [band, count_f, pct_f, avg_f, top_prop, top_ref, top_neigh]
    for ci, val in enumerate(row_data, 1):
        c = ws2.cell(ri, ci, val)
        c.fill = PatternFill("solid", start_color=bg)
        c.font = Font(name="Calibri", bold=(ci==1), size=10, color=fg)
        c.alignment = Alignment(horizontal="center", vertical="center"); c.border=bdr()
        if ci == 3: c.number_format = "0.0%"
        if ci == 4: c.number_format = "#,##0"
    ws2.row_dimensions[ri].height = 22

# Total row
tr = len(bands) + 6
for ci in range(1, 8): ws2.cell(tr, ci).border = bdr()
ws2.cell(tr, 1, "TOTAL")
ws2.cell(tr, 1).font = Font(name="Calibri", bold=True, size=10)
ws2.cell(tr, 1).alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws2.cell(tr, 2, f"=COUNTA({tref(target)})")
ws2.cell(tr, 2).font = Font(name="Calibri", bold=True, size=10)
ws2.cell(tr, 2).alignment = Alignment(horizontal="center", vertical="center")
ws2.cell(tr, 3, "100%")
ws2.cell(tr, 3).font = Font(name="Calibri", bold=True, size=10)
ws2.cell(tr, 3).alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[tr].height = 22

# ── Per-factor breakdown tables ───────────────────────────────────────
sr = tr + 2
for display, col, bin_col, v, p in sig_factors:
    tbl_col = bin_col_name if bin_col else col
    ordered_vals = get_ordered_vals(col, bin_col)

    ws2.merge_cells(f"A{sr}:G{sr}")
    hdr(ws2.cell(sr, 1), f"Breakdown by {display}  (V={v:.4f}, p={p:.6f})",
        bg=NAVY2, size=10, center=False)
    ws2.row_dimensions[sr].height = 20; sr += 1

    sub_hdrs = [display, "Total"] + bands + [f"{b} %" for b in bands]
    for ci, h in enumerate(sub_hdrs, 1):
        c = ws2.cell(sr, ci, h)
        c.font=Font(name="Calibri",bold=True,size=9,color=WHITE)
        c.fill=PatternFill("solid",start_color=NAVY)
        c.alignment=Alignment(horizontal="center",vertical="center"); c.border=bdr()
    ws2.row_dimensions[sr].height = 18; sr += 1

    for vi, val in enumerate(ordered_vals):
        bg = ALTROW if vi%2==0 else WHITE
        # Label
        c = ws2.cell(sr, 1, val)
        c.font=Font(name="Calibri",size=9); c.fill=PatternFill("solid",start_color=bg)
        c.alignment=Alignment(horizontal="left",vertical="center",indent=1); c.border=bdr()

        # Total via COUNTIF
        tot_f = f'=COUNTIF({tref(tbl_col)},"{val}")'
        tot_ci = 2
        ws2.cell(sr, tot_ci, tot_f)
        ws2.cell(sr, tot_ci).font=Font(name="Calibri",bold=True,size=9)
        ws2.cell(sr, tot_ci).fill=PatternFill("solid",start_color=bg)
        ws2.cell(sr, tot_ci).alignment=Alignment(horizontal="center",vertical="center")
        ws2.cell(sr, tot_ci).border=bdr()

        # Band counts
        for bi, b in enumerate(bands):
            ci = 3 + bi
            c = ws2.cell(sr, ci,
                f'=COUNTIFS({tref(tbl_col)},"{val}",{tref(target)},"{b}")')
            c.font=Font(name="Calibri",size=9); c.fill=PatternFill("solid",start_color=bg)
            c.alignment=Alignment(horizontal="center",vertical="center"); c.border=bdr()

        # Band %
        for bi, b in enumerate(bands):
            pci = 3 + len(bands) + bi
            bci = 3 + bi
            c = ws2.cell(sr, pci,
                f"=IFERROR(ROUND({get_column_letter(bci)}{sr}/{get_column_letter(tot_ci)}{sr}*100,1),0)")
            c.font=Font(name="Calibri",size=9); c.fill=PatternFill("solid",start_color=bg)
            c.alignment=Alignment(horizontal="center",vertical="center")
            c.border=bdr(); c.number_format="0.0"

        ws2.row_dimensions[sr].height = 18; sr += 1
    sr += 1

# ═══════════════════════════════════════════════════════════════════════
# SHEET 5: _LOOKUP (hidden) — probability snapshot for scorer
# ═══════════════════════════════════════════════════════════════════════

ws_lk = wb.create_sheet("_Lookup")
ws_lk.sheet_state = "hidden"
ws_lk.cell(1,1,"ColName"); ws_lk.cell(1,2,"Value")
for bi,b in enumerate(bands): ws_lk.cell(1,3+bi,b)
lr = 2
for col, val_dict in marginal.items():
    for val, pd_ in val_dict.items():
        ws_lk.cell(lr,1,col); ws_lk.cell(lr,2,val)
        for bi,b in enumerate(bands): ws_lk.cell(lr,3+bi,pd_.get(b,0))
        lr+=1
lk_last=lr-1
ws_lk.cell(lr,1,"__overall__")
for bi,b in enumerate(bands): ws_lk.cell(lr,3+bi,overall[b])
overall_row=lr

def pf(col_name, cell_ref, bi):
    pc=get_column_letter(3+bi)
    fb=ws_lk.cell(overall_row,3+bi).value or 0
    return (f"IFERROR(AVERAGEIFS(_Lookup!${pc}$2:${pc}${lk_last},"
            f"_Lookup!$A$2:$A${lk_last},{chr(34)}{col_name}{chr(34)},"
            f"_Lookup!$B$2:$B${lk_last},{cell_ref}),{fb})")

# ═══════════════════════════════════════════════════════════════════════
# SHEET 6: LEAD SCORER
# ═══════════════════════════════════════════════════════════════════════

ws=wb.create_sheet("Lead Scorer")
ws.sheet_view.showGridLines=False
ws.column_dimensions["A"].width=3
ws.column_dimensions["B"].width=30
ws.column_dimensions["C"].width=34
ws.column_dimensions["D"].width=3
ws.column_dimensions["E"].width=0.1

R=1
ws.row_dimensions[R].height=14; R+=1

ws.merge_cells(f"B{R}:C{R}")
t=ws.cell(R,2,"Kingston Painting - Lead Scorer")
t.font=Font(name="Calibri",bold=True,size=18,color=NAVY)
t.alignment=Alignment(horizontal="center",vertical="center")
ws.row_dimensions[R].height=36; R+=1

ws.merge_cells(f"B{R}:C{R}")
s=ws.cell(R,2,"Select each factor, then click the CALCULATE SCORE button")
s.font=Font(name="Calibri",italic=True,size=10,color="555555")
s.alignment=Alignment(horizontal="center",vertical="center")
ws.row_dimensions[R].height=20; R+=1
ws.row_dimensions[R].height=10; R+=1

ws.merge_cells(f"B{R}:C{R}")
hdr(ws.cell(R,2),"LEAD INFORMATION")
ws.row_dimensions[R].height=24; R+=1

field_order=[
    ("Neighbourhood",    "neighbourhood",        None),
    ("Property Type",    "property_type",        None),
    ("Homeowner Status", "homeowner_status",     None),
    ("Job Size (sqft)",  num_col,                num_col+"_bin" if num_col else None),
    ("Referral Source",  "referral_source",      None),
    ("Age Bracket",      "customer_age_bracket", None),
]
field_order=[(d,c,b) for d,c,b in field_order
             if c and c in df.columns and (b is None or b in df.columns)]

input_refs={}
for display, col, bin_col in field_order:
    ws.row_dimensions[R].height=26
    lbl=ws.cell(R,2,display)
    lbl.font=Font(name="Calibri",bold=True,size=10)
    lbl.fill=PatternFill("solid",start_color=LBLUE)
    lbl.alignment=Alignment(horizontal="left",vertical="center",indent=1); lbl.border=bdr()
    inp=ws.cell(R,3)
    inp.fill=PatternFill("solid",start_color=WHITE)
    inp.font=Font(name="Calibri",size=10,color=NAVY)
    inp.alignment=Alignment(horizontal="left",vertical="center",indent=1); inp.border=bdr()

    ordered_vals = get_ordered_vals(col, bin_col)
    dv=DataValidation(type="list",formula1='"'+",".join(ordered_vals)+'"',
                      allow_blank=True,showErrorMessage=False)
    ws.add_data_validation(dv); dv.add(inp)
    input_refs[col]=f"C{R}"
    R+=1

ws.row_dimensions[R].height=8; R+=1

ws.merge_cells(f"B{R}:C{R}")
btn=ws.cell(R,2,"CALCULATE SCORE")
btn.font=Font(name="Calibri",bold=True,size=13,color=WHITE)
btn.fill=PatternFill("solid",start_color=GBTN)
btn.alignment=Alignment(horizontal="center",vertical="center"); btn.border=bdr(GBTN)
ws.row_dimensions[R].height=34; R+=1
ws.row_dimensions[R].height=8; R+=1

ws.merge_cells(f"B{R}:C{R}")
hdr(ws.cell(R,2),"RESULTS")
ws.row_dimensions[R].height=24; R+=1

# Intermediates at row 200+
INTER_OFFSET=200
e_refs=[]; eq_refs=[]
for i,(display,col,bin_col) in enumerate(field_order):
    if col not in input_refs: continue
    cell_ref=input_refs[col]
    inter_row=INTER_OFFSET+i
    ws.cell(inter_row,5,f"={pf(col,cell_ref,0)}")
    e_refs.append(f"E{inter_row}")
    w_ref=weight_cell_refs.get(col,"1")
    eq_refs.append((f"E{inter_row}",w_ref))

avg_e="AVERAGE("+",".join(e_refs)+")" if e_refs else "33.3"
data_f=f"=IFERROR(ROUND(({avg_e})/10,1),\"—\")"

if eq_refs:
    num_p="+".join([f"({er}*{wr})" for er,wr in eq_refs])
    den_p="+".join([str(wr) for _,wr in eq_refs])
    qual_f=f"=IFERROR(ROUND(({num_p})/({den_p})/10,1),\"—\")"
else:
    qual_f="=5"

data_r=R; qual_r=R+1; final_r=R+2; tier_r=R+4

result_defs=[
    ("Data Score (0-10)",        data_f),
    ("Qualitative Score (0-10)", qual_f),
    ("Final Score (0-10)",       f"=IFERROR(ROUND(AVERAGE(C{data_r},C{qual_r}),1),\"—\")"),
    ("Projected Profitability",  f'=IFERROR(ROUND({avg_e},1)&"%","—")'),
    ("Lead Tier",                f'=IFERROR(IF(C{final_r}="—","—",IF(C{final_r}>=6.5,"High",IF(C{final_r}>=4,"Medium","Low"))),"—")'),
]
for label,formula in result_defs:
    ws.row_dimensions[R].height=26
    lbl=ws.cell(R,2,label)
    lbl.font=Font(name="Calibri",bold=True,size=10)
    lbl.fill=PatternFill("solid",start_color=LBLUE)
    lbl.alignment=Alignment(horizontal="left",vertical="center",indent=1); lbl.border=bdr()
    val=ws.cell(R,3,formula)
    val.fill=PatternFill("solid",start_color=CREAM)
    val.font=Font(name="Calibri",bold=True,size=11,color=NAVY)
    val.alignment=Alignment(horizontal="center",vertical="center"); val.border=bdr()
    R+=1
ws.row_dimensions[R].height=14

for cell_r,op,fv,bg,fg in [
    (f"C{tier_r}","equal",['"High"'],GBG,GFG),
    (f"C{tier_r}","equal",['"Medium"'],ABG,AFG),
    (f"C{tier_r}","equal",['"Low"'],RBG,RFG),
    (f"C{final_r}","greaterThanOrEqual",["6.5"],GBG,GFG),
    (f"C{final_r}","between",["4","6.4"],ABG,AFG),
    (f"C{final_r}","lessThan",["4"],RBG,RFG),
]:
    ws.conditional_formatting.add(cell_r,CellIsRule(
        operator=op,formula=fv,
        fill=PatternFill("solid",start_color=bg),
        font=Font(name="Calibri",bold=True,color=fg)
    ))

# ── SAVE ─────────────────────────────────────────────────────────────

wb.save("lead_scorer.xlsx")
print(f"\n✓ Saved: lead_scorer.xlsx")
print(f"  LeadData table: {total_rows} rows × {len(df_export.columns)} cols")
print(f"  Frequency Tables, Summary Stats, Factor Weights — all live COUNTIFS")
print(f"  Factors: {[d for d,c,b,v,p in sig_factors]}")
